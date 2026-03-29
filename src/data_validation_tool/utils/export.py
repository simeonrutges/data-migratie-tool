import pandas as pd
from openpyxl.styles import Font, PatternFill


def export_differences_to_csv(differences: list[dict], output_path: str) -> None:
    """
    Exporteert veldverschillen naar een CSV-bestand.

    Parameters:
    - differences: lijst met verschillen (dicts met keys: id, kolom, bron, doel)
    - output_path: pad naar het outputbestand
    """
    if not differences:
        df = pd.DataFrame(columns=["id", "kolom", "bron", "doel"])
    else:
        df = pd.DataFrame(differences)

    df.to_csv(output_path, index=False)


def export_summary_to_csv(summary_data: list[dict], output_path: str) -> None:
    """
    Exporteert samenvattingsgegevens naar een CSV-bestand.

    Parameters:
    - summary_data: lijst met samenvattingsregels
    - output_path: pad naar het outputbestand
    """
    df = pd.DataFrame(summary_data)
    df.to_csv(output_path, index=False)


def export_distribution_to_csv(distribution_data: list[dict], output_path: str) -> None:
    """
    Exporteert distributieverschillen naar een CSV-bestand.

    Parameters:
    - distribution_data: lijst met distributieverschillen per kolom/waarde
    - output_path: pad naar het outputbestand
    """
    df = pd.DataFrame(distribution_data)
    df.to_csv(output_path, index=False)


from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def style_worksheet(worksheet) -> None:
    """
    Past eenvoudige, leesbare styling toe op een Excel-worksheet.

    Styling:
    - modern lettertype (Aptos, 14 pt)
    - header vet + achtergrondkleur
    - autofilter op eerste rij
    - eerste rij vastzetten
    - kolombreedtes ruimer instellen op basis van inhoud
    """
    header_fill = PatternFill(
        fill_type="solid",
        start_color="BDD7EE",
        end_color="BDD7EE",
    )

    header_font = Font(name="Aptos", size=14, bold=True)
    body_font = Font(name="Aptos", size=14)

    # Style de volledige sheet
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = body_font

    # Style de header-rij apart
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Zet een filter op de header-rij
    worksheet.auto_filter.ref = worksheet.dimensions

    # Zorg dat de header zichtbaar blijft bij scrollen
    worksheet.freeze_panes = "A2"

    # Pas kolombreedte aan op basis van de langste waarde
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0

        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)

            # Iets ruimere schatting:
            # langere teksten krijgen iets meer marge
            adjusted_length = len(cell_value)

            if adjusted_length > max_length:
                max_length = adjusted_length

        # Extra marge voor leesbaarheid
        # Min breedte voorkomt te smalle kolommen
        calculated_width = max(max_length + 4, 14)
        worksheet.column_dimensions[column_letter].width = calculated_width


def highlight_differences(worksheet) -> None:
    """
    Highlight de 'bron' en 'doel' kolommen in de field_differences sheet.

    Doel:
    - visueel duidelijk maken waar verschillen zitten
    """

    highlight_fill = PatternFill(
        fill_type="solid",
        start_color="FFF2CC",
        end_color="FFF2CC",
    )

    # Zoek kolomindexen op basis van header
    header = [cell.value for cell in worksheet[1]]

    try:
        bron_index = header.index("bron") + 1
        doel_index = header.index("doel") + 1
    except ValueError:
        # Als kolommen niet bestaan, doe niets
        return

    # Loop door alle rijen behalve header
    for row in worksheet.iter_rows(min_row=2):
        row[bron_index - 1].fill = highlight_fill
        row[doel_index - 1].fill = highlight_fill


def export_report_to_excel(
    summary_data: list[dict],
    distribution_data: list[dict],
    differences_data: list[dict],
    output_path: str,
) -> None:
    """
    Exporteert het volledige validatierapport naar één Excel-bestand.

    Sheets:
    - summary
    - distribution
    - field_differences

    Parameters:
    - summary_data: samenvattingsgegevens
    - distribution_data: distributieverschillen
    - differences_data: veldverschillen
    - output_path: pad naar het Excel-bestand
    """
    summary_df = pd.DataFrame(summary_data)
    distribution_df = pd.DataFrame(distribution_data)

    if not differences_data:
        differences_df = pd.DataFrame(columns=["id", "kolom", "bron", "doel"])
    else:
        differences_df = pd.DataFrame(differences_data)

    # Schrijf alle dataframes naar aparte tabbladen in één workbook
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        distribution_df.to_excel(writer, sheet_name="distribution", index=False)
        differences_df.to_excel(
            writer,
            sheet_name="field_differences",
            index=False,
        )

        # Haal worksheets op en geef ze consistente styling
        workbook = writer.book
        style_worksheet(workbook["summary"])
        style_worksheet(workbook["distribution"])
        style_worksheet(workbook["field_differences"])

        # Highlight verschillen in de field_differences sheet
        highlight_differences(workbook["field_differences"])
